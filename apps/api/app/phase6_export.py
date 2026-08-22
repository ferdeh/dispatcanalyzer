from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from .phase6_service import get_prediction_run


def workbook_bytes(sheets: list[tuple[str, list[str], list[list]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, headers, rows in sheets:
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B73BF")
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(50, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def loading_order_template() -> bytes:
    return workbook_bytes(
        [("Loading Order", ["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"], [["LO000001", "2026-08-22 05:00:00", "SPBU001", 8]])]
    )


def mt_availability_template() -> bytes:
    return workbook_bytes(
        [("MT Availability", ["vehicle_registration_no", "initial_available_datetime"], [["B9123ABC", "2026-08-22 04:30:00"]])]
    )


def validation_report(issues: list[dict]) -> bytes:
    headers = ["File", "Row", "Field", "Status", "Error Code", "Description"]
    rows = [[issue["file"], issue["row"], issue["field"], issue["status"], issue["error_code"], issue["description"]] for issue in issues]
    return workbook_bytes([("Validation", headers, rows)])


def prediction_export(db: Session, run_id: str) -> tuple[bytes, str]:
    payload = get_prediction_run(db, run_id)
    summary_rows = [
        ["prediction_run_id", payload["prediction_run_id"]],
        ["depot", payload["depot"]],
        ["model", payload["model"].get("model_name")],
        ["model_version", payload["model"].get("model_version")],
        ["run_datetime", payload["created_at"]],
        ["user", payload["created_by"]],
        *[[key, value] for key, value in payload["summary"].items()],
    ]
    shipment_rows = []
    assignment_rows = []
    candidate_rows = []
    trip_rows = []
    for shipment in payload["shipments"]:
        trip = shipment.get("trip") or {}
        for line in shipment["lines"]:
            shipment_rows.append(
                [
                    payload["prediction_run_id"],
                    shipment["shift"],
                    shipment["planned_start_datetime"],
                    shipment["predicted_shipment_id"],
                    line["loading_order_no"],
                    line["shipment_start_datetime"],
                    line["spbu_no"],
                    line["order_quantity_kl"],
                    shipment["shipment_prediction_score"],
                    shipment["shipment_confidence_level"],
                ]
            )
        assignment = shipment["assignment"]
        assignment_rows.append(
            [
                payload["prediction_run_id"],
                shipment["shift"],
                shipment["predicted_shipment_id"],
                assignment["assigned_vehicle_registration"],
                assignment["mt_assignment_score"],
                assignment["assignment_status"],
                shipment["is_manual_override"] or assignment["assignment_status"] == "MANUAL_OVERRIDE",
                assignment["unassigned_reason"],
            ]
        )
        for candidate in shipment["candidates"]:
            candidate_rows.append(
                [
                    shipment["predicted_shipment_id"],
                    candidate["vehicle_registration_no"],
                    candidate["candidate_rank"],
                    candidate["prediction_score"],
                    candidate["compatibility_status"],
                    candidate["exclusion_reason"],
                ]
            )
        trip_rows.append(
            [
                trip.get("trip_id"),
                trip.get("trip_number"),
                trip.get("vehicle_registration_no"),
                shipment["predicted_shipment_id"],
                trip.get("planned_start_datetime"),
                trip.get("predicted_departure_datetime"),
                trip.get("delay_minutes"),
                ", ".join(trip.get("estimated_visit_sequence") or []),
                trip.get("routing_provider"),
                trip.get("routing_mode"),
                trip.get("route_distance_meters"),
                trip.get("route_duration_seconds"),
                trip.get("service_duration_seconds"),
                trip.get("total_cycle_duration_seconds"),
                trip.get("estimated_return_datetime"),
                trip.get("next_available_datetime"),
                trip.get("routing_confidence"),
                trip.get("route_estimation_source"),
                trip.get("assignment_status"),
                trip.get("fallback_used"),
                ", ".join(trip.get("warning_codes") or []),
            ]
        )
    validation_rows = [
        [issue["file"], issue["row"], issue["field"], issue["status"], issue["error_code"], issue["description"]]
        for issue in payload["validation"]
    ]
    content = workbook_bytes(
        [
            ("Summary", ["Metric", "Value"], summary_rows),
            (
                "Shipment Result",
                ["prediction_run_id", "derived_shift", "planned_start_datetime", "predicted_shipment_id", "loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl", "shipment_prediction_score", "shipment_confidence_level"],
                shipment_rows,
            ),
            (
                "Trip Timeline",
                ["trip_id", "trip_number", "vehicle_registration_no", "predicted_shipment_id", "planned_start_datetime", "predicted_departure_datetime", "delay_minutes", "estimated_visit_sequence", "routing_provider", "routing_mode", "route_distance_meters", "route_duration_seconds", "service_duration_seconds", "total_cycle_duration_seconds", "estimated_return_datetime", "next_available_datetime", "routing_confidence", "route_estimation_source", "assignment_status", "fallback_used", "warning_codes"],
                trip_rows,
            ),
            (
                "MT Assignment",
                ["prediction_run_id", "shift", "predicted_shipment_id", "vehicle_registration_no", "mt_assignment_score", "assignment_status", "override_status", "unassigned_reason"],
                assignment_rows,
            ),
            (
                "MT Candidates",
                ["predicted_shipment_id", "vehicle_registration_no", "candidate_rank", "prediction_score", "compatibility_status", "exclusion_reason"],
                candidate_rows,
            ),
            ("Validation", ["file", "row", "field", "status", "error_code", "description"], validation_rows),
        ]
    )
    return content, f"{payload['prediction_run_id']}-prediction-result.xlsx"
