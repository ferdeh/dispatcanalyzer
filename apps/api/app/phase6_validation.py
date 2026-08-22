from __future__ import annotations

from collections import Counter
from datetime import datetime, timedelta, timezone
from io import BytesIO
from math import isfinite
from time import perf_counter
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .departure_intelligence import shift_for_minute, validate_shift_config
from .models import MLBehavioralModel, MasterDepot, MasterMT, MasterSPBU, SpbuIdentifierAlias
from .normalization import clean_str, mt_capacity_kl, normalize_key


MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
LOADING_ORDER_COLUMNS = ("loading_order_no", "shipment_start_datetime", "spbu_no")
MT_AVAILABILITY_COLUMNS = ("vehicle_registration_no", "initial_available_datetime")


def require_prediction_model(db: Session, depot_id: str, model_id: str) -> MLBehavioralModel:
    model = db.get(MLBehavioralModel, model_id)
    if not model:
        raise HTTPException(status_code=404, detail={"code": "MODEL_NOT_FOUND", "message": "Prediction model was not found."})
    if model.depot_id != depot_id:
        raise HTTPException(status_code=409, detail={"code": "MODEL_DEPOT_MISMATCH", "message": "Prediction model belongs to another depot."})
    if model.model_status not in {"SAVED", "ACTIVE", "READY"}:
        raise HTTPException(status_code=409, detail={"code": "MODEL_NOT_READY", "message": "Prediction model is not ready for inference."})
    return model


def _issue(file_name: str, row: int, field: str, status: str, code: str, description: str) -> dict:
    return {"file": file_name, "row": row, "field": field, "status": status, "error_code": code, "description": description}


def read_excel_rows(
    content: bytes,
    *,
    file_name: str,
    required_columns: tuple[str, ...],
    aliases: dict[str, str] | None = None,
) -> tuple[list[dict], list[dict]]:
    if not content or len(content) > MAX_WORKBOOK_BYTES:
        return [], [_issue(file_name, 1, "file", "ERROR", "INVALID_FILE_SIZE", "Workbook must be non-empty and no larger than 10 MB.")]
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        values = workbook.active.iter_rows(values_only=True)
        raw_headers = next(values, None)
        if not raw_headers:
            return [], [_issue(file_name, 1, "header", "ERROR", "REQUIRED_COLUMNS_MISSING", "Workbook has no header row.")]
        alias_map = {normalize_key(key): value for key, value in (aliases or {}).items()}
        headers = []
        for value in raw_headers:
            raw = clean_str(value) or ""
            headers.append(alias_map.get(normalize_key(raw) or "", raw.strip().lower()))
        missing = [column for column in required_columns if column not in headers]
        if missing:
            return [], [_issue(file_name, 1, column, "ERROR", "REQUIRED_COLUMN_MISSING", f"Required column '{column}' is missing.") for column in missing]
        rows = []
        for row_number, raw_row in enumerate(values, start=2):
            row = {headers[index]: clean_str(value) for index, value in enumerate(raw_row) if index < len(headers) and headers[index]}
            if any(value is not None for value in row.values()):
                row["_row_number"] = row_number
                rows.append(row)
        return rows, []
    except Exception:
        return [], [_issue(file_name, 1, "file", "ERROR", "INVALID_EXCEL_FILE", "File could not be read as an Excel workbook.")]


def shift_lookup(model: MLBehavioralModel) -> dict[str, dict]:
    """Compatibility helper used by demo generation and older callers."""
    lookup: dict[str, dict] = {}
    for shift in model.shift_definition_snapshot or []:
        cleaned = {
            "shift_id": clean_str(shift.get("shift_id")) or "",
            "name": clean_str(shift.get("name")) or clean_str(shift.get("shift_id")) or "",
            "start_time": clean_str(shift.get("start_time")),
            "end_time": clean_str(shift.get("end_time")),
        }
        for value in (cleaned["shift_id"], cleaned["name"]):
            if normalize_key(value):
                lookup[normalize_key(value)] = cleaned
    return lookup


def _model_shift_config(model: MLBehavioralModel) -> list[dict]:
    try:
        return validate_shift_config(model.shift_definition_snapshot or [])
    except HTTPException as exc:
        raise HTTPException(
            status_code=409,
            detail={"code": "MODEL_SHIFT_CONFIGURATION_INVALID", "message": f"Selected Phase 5 model has an invalid shift snapshot: {exc.detail}"},
        ) from exc


def _depot_timezone(db: Session, depot_id: str) -> ZoneInfo:
    depot = db.get(MasterDepot, depot_id)
    name = depot.timezone if depot and depot.timezone else "Asia/Jakarta"
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError as exc:
        raise HTTPException(status_code=409, detail={"code": "DEPOT_TIMEZONE_INVALID", "message": f"Depot timezone '{name}' is invalid."}) from exc


def _parse_datetime(value: Any, depot_timezone: ZoneInfo) -> datetime | None:
    raw = clean_str(value)
    if not raw:
        return None
    normalized = raw.strip().replace("Z", "+00:00")
    parsed: datetime | None = None
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        for pattern in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                parsed = datetime.strptime(raw, pattern)
                break
            except ValueError:
                continue
    if parsed is None:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=depot_timezone)
    return parsed.astimezone(timezone.utc)


def _order_quantity_kl(value: Any) -> tuple[float | None, bool]:
    raw = clean_str(value)
    if raw is None:
        return None, True
    try:
        quantity = float(raw.replace(",", "."))
    except ValueError:
        return None, False
    if not isfinite(quantity) or quantity <= 0:
        return None, False
    return round(quantity, 3), True


def validate_loading_orders(
    db: Session,
    *,
    depot_id: str,
    model: MLBehavioralModel,
    content: bytes,
    file_name: str,
    maximum_planning_horizon_days: int = 7,
) -> dict:
    started = perf_counter()
    rows, issues = read_excel_rows(
        content,
        file_name=file_name,
        required_columns=LOADING_ORDER_COLUMNS,
        aliases={
            "loading_order_no": "loading_order_no",
            "loading_order_number": "loading_order_no",
            "loading_order_id": "loading_order_no",
            "shipment_start_datetime": "shipment_start_datetime",
            "planned_start_datetime": "shipment_start_datetime",
            "start_datetime": "shipment_start_datetime",
            "spbu_no": "spbu_no",
            "spbu_code": "spbu_no",
            "spbu_id": "spbu_no",
            "order_quantity_kl": "order_quantity_kl",
            "quantity_kl": "order_quantity_kl",
            "total_order_kl": "order_quantity_kl",
            "volume_kl": "order_quantity_kl",
        },
    )
    if issues:
        return _validation_payload("LOADING_ORDER", [], issues, started)

    timezone_info = _depot_timezone(db, depot_id)
    shift_config = _model_shift_config(model)
    spbus = db.scalars(select(MasterSPBU)).all()
    spbu_by_id = {spbu.spbu_id: spbu for spbu in spbus}
    spbu_by_key = {key: spbu for spbu in spbus for key in {normalize_key(spbu.spbu_id), normalize_key(spbu.spbu_code)} if key}
    for alias in db.scalars(select(SpbuIdentifierAlias).where(SpbuIdentifierAlias.active_status == "ACTIVE")).all():
        if normalize_key(alias.identifier_value) and alias.spbu_id in spbu_by_id:
            spbu_by_key[normalize_key(alias.identifier_value)] = spbu_by_id[alias.spbu_id]
    duplicates = Counter(normalize_key(row.get("loading_order_no")) for row in rows if normalize_key(row.get("loading_order_no")))
    normalized_rows = []
    for row in rows:
        row_number = int(row["_row_number"])
        for field in LOADING_ORDER_COLUMNS:
            if not clean_str(row.get(field)):
                issues.append(_issue(file_name, row_number, field, "ERROR", "REQUIRED_VALUE_EMPTY", f"{field} must not be empty."))
        lo_key = normalize_key(row.get("loading_order_no"))
        if lo_key and duplicates[lo_key] > 1:
            issues.append(_issue(file_name, row_number, "loading_order_no", "ERROR", "DUPLICATE_LOADING_ORDER", "Loading Order appears more than once."))
        spbu = spbu_by_key.get(normalize_key(row.get("spbu_no")) or "")
        if row.get("spbu_no") and not spbu:
            issues.append(_issue(file_name, row_number, "spbu_no", "ERROR", "SPBU_NOT_FOUND", "SPBU identifier is not present in canonical master data."))
        elif spbu and spbu.primary_depot_id != depot_id:
            issues.append(_issue(file_name, row_number, "spbu_no", "ERROR", "SPBU_DEPOT_MISMATCH", "SPBU belongs to another depot."))
        parsed = _parse_datetime(row.get("shipment_start_datetime"), timezone_info)
        if row.get("shipment_start_datetime") and not parsed:
            issues.append(_issue(file_name, row_number, "shipment_start_datetime", "ERROR", "INVALID_DATETIME", "shipment_start_datetime must be a valid complete datetime."))
        quantity, quantity_valid = _order_quantity_kl(row.get("order_quantity_kl"))
        if not quantity_valid:
            issues.append(_issue(file_name, row_number, "order_quantity_kl", "ERROR", "INVALID_ORDER_QUANTITY", "order_quantity_kl must be greater than 0 when provided."))
        if lo_key and duplicates[lo_key] == 1 and spbu and spbu.primary_depot_id == depot_id and parsed and quantity_valid:
            local = parsed.astimezone(timezone_info)
            shift = shift_for_minute(local.hour * 60 + local.minute, shift_config)
            normalized_rows.append(
                {
                    "source_row_number": row_number,
                    "loading_order_no": clean_str(row["loading_order_no"]),
                    "shipment_start_datetime": parsed.isoformat(),
                    "shipment_start_datetime_local": local.isoformat(),
                    "shift_id": shift["shift_id"],
                    "shift": shift["name"],
                    "spbu_id": spbu.spbu_id,
                    "spbu_no": spbu.spbu_code,
                    "spbu_name": spbu.spbu_name,
                    "order_quantity_kl": quantity,
                }
            )
    datetimes = [datetime.fromisoformat(row["shipment_start_datetime"]) for row in normalized_rows]
    if datetimes and max(datetimes) - min(datetimes) > timedelta(days=max(1, maximum_planning_horizon_days)):
        issues.append(_issue(file_name, 1, "shipment_start_datetime", "ERROR", "PLANNING_HORIZON_EXCEEDED", f"Loading Orders must fit within a {maximum_planning_horizon_days}-day planning horizon."))
    return _validation_payload("LOADING_ORDER", normalized_rows, issues, started)


def validate_mt_availability(db: Session, *, depot_id: str, model: MLBehavioralModel, content: bytes, file_name: str) -> dict:
    started = perf_counter()
    rows, issues = read_excel_rows(
        content,
        file_name=file_name,
        required_columns=MT_AVAILABILITY_COLUMNS,
        aliases={
            "vehicle_registration_no": "vehicle_registration_no",
            "vehicle_registration": "vehicle_registration_no",
            "registration_no": "vehicle_registration_no",
            "mt_id": "vehicle_registration_no",
            "initial_available_datetime": "initial_available_datetime",
            "available_datetime": "initial_available_datetime",
            "next_available_datetime": "initial_available_datetime",
        },
    )
    if issues:
        return _validation_payload("MT_AVAILABILITY", [], issues, started)
    timezone_info = _depot_timezone(db, depot_id)
    mts = db.scalars(select(MasterMT)).all()
    mt_by_key = {key: mt for mt in mts for key in {normalize_key(mt.mt_id), normalize_key(mt.vehicle_registration), normalize_key(mt.source_mt_id)} if key}
    duplicates = Counter(normalize_key(row.get("vehicle_registration_no")) for row in rows if normalize_key(row.get("vehicle_registration_no")))
    normalized_rows = []
    for row in rows:
        row_number = int(row["_row_number"])
        for field in MT_AVAILABILITY_COLUMNS:
            if not clean_str(row.get(field)):
                issues.append(_issue(file_name, row_number, field, "ERROR", "REQUIRED_VALUE_EMPTY", f"{field} must not be empty."))
        vehicle_key = normalize_key(row.get("vehicle_registration_no"))
        if vehicle_key and duplicates[vehicle_key] > 1:
            issues.append(_issue(file_name, row_number, "vehicle_registration_no", "ERROR", "DUPLICATE_VEHICLE_AVAILABILITY", "Vehicle must have exactly one initial availability record."))
        mt = mt_by_key.get(vehicle_key or "")
        if row.get("vehicle_registration_no") and not mt:
            issues.append(_issue(file_name, row_number, "vehicle_registration_no", "ERROR", "VEHICLE_NOT_FOUND", "Vehicle is not present in canonical MT master data."))
        elif mt and mt.depot_id and mt.depot_id != depot_id:
            issues.append(_issue(file_name, row_number, "vehicle_registration_no", "ERROR", "VEHICLE_DEPOT_MISMATCH", "Vehicle belongs to another depot."))
        elif mt and mt.active_status != "ACTIVE":
            issues.append(_issue(file_name, row_number, "vehicle_registration_no", "ERROR", "VEHICLE_INACTIVE", "Vehicle master is not active."))
        parsed = _parse_datetime(row.get("initial_available_datetime"), timezone_info)
        if row.get("initial_available_datetime") and not parsed:
            issues.append(_issue(file_name, row_number, "initial_available_datetime", "ERROR", "INVALID_AVAILABLE_DATETIME", "initial_available_datetime must be a valid complete datetime."))
        if vehicle_key and duplicates[vehicle_key] == 1 and mt and (not mt.depot_id or mt.depot_id == depot_id) and mt.active_status == "ACTIVE" and parsed:
            normalized_rows.append(
                {
                    "source_row_number": row_number,
                    "vehicle_id": mt.mt_id,
                    "vehicle_registration_no": mt.vehicle_registration or mt.mt_id,
                    "capacity_kl": mt_capacity_kl(mt.capacity_label, mt.vehicle_type_tag),
                    "initial_available_datetime": parsed.isoformat(),
                    "initial_available_datetime_local": parsed.astimezone(timezone_info).isoformat(),
                }
            )
    return _validation_payload("MT_AVAILABILITY", normalized_rows, issues, started)


def _validation_payload(file_type: str, rows: list[dict], issues: list[dict], started: float) -> dict:
    errors = sum(issue["status"] == "ERROR" for issue in issues)
    warnings = sum(issue["status"] == "WARNING" for issue in issues)
    return {
        "file_type": file_type,
        "status": "ERROR" if errors else "WARNING" if warnings else "PASS",
        "blocking_error_count": errors,
        "warning_count": warnings,
        "issues": issues,
        "normalized_rows": rows,
        "row_count": len(rows),
        "detected_shifts": sorted({row["shift"] for row in rows if row.get("shift")}),
        "duration_ms": round((perf_counter() - started) * 1000),
    }
