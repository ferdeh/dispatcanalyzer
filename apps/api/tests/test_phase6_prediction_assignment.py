from __future__ import annotations

import json
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

import httpx
import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.config import get_settings
from app.database import get_db
from app.google_routes import (
    GoogleRoutesClient,
    GoogleRoutesError,
    get_google_routes_configuration,
    public_google_routes_configuration,
    save_google_routes_configuration,
)
from app.models import (
    Base,
    BridgeMTTag,
    BridgeSPBUTag,
    MLBehavioralModel,
    MLSPBUClusterAssignment,
    MasterDepot,
    MasterMT,
    MasterProduct,
    MasterSPBU,
    MasterTag,
    MasterTagType,
    PredictionJob,
)
from app.main import app
from app.phase6_demo import generate_demo_loading_orders, generate_demo_mt_availability
from app.phase6_export import loading_order_workbook
from app.phase6_routing import Phase6RouteEstimationService
from app.phase6_jobs import claim_next_prediction_job, heartbeat_prediction_job, recover_stale_prediction_jobs, utc_now
from app.phase6_service import (
    create_prediction_run,
    enqueue_prediction_run,
    get_prediction_run,
    get_prediction_shipment_candidates,
    get_prediction_run_status,
    override_assignment,
    process_prediction_run,
)
from app.phase6_validation import validate_loading_orders, validate_mt_availability


def make_session():
    engine = create_engine("sqlite+pysqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)

    @event.listens_for(engine, "connect")
    def enable_sqlite_foreign_keys(dbapi_connection, _connection_record) -> None:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)


def workbook(headers: list[str], rows: list[list]) -> bytes:
    result = Workbook()
    sheet = result.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    result.save(buffer)
    return buffer.getvalue()


def seed(session) -> MLBehavioralModel:
    session.add_all(
        [
            MasterDepot(depot_id="D1", depot_code="D1", depot_name="Depot One", latitude=-6.20, longitude=106.84, timezone="Asia/Jakarta"),
            MasterDepot(depot_id="D2", depot_code="D2", depot_name="Depot Two", timezone="Asia/Jakarta"),
        ]
    )
    session.commit()
    session.add_all(
        [
            MasterMT(mt_id="T1", vehicle_name_raw="Truck 1", vehicle_registration="B1001AA", capacity_label="8KL", vehicle_type_tag=8, number_of_compartments=1, depot_id="D1", active_status="ACTIVE"),
            MasterMT(mt_id="T2", vehicle_name_raw="Truck 2", vehicle_registration="B1002AA", capacity_label="8 KL", vehicle_type_tag=8, number_of_compartments=1, depot_id="D1", active_status="ACTIVE"),
            MasterMT(mt_id="T3", vehicle_name_raw="Truck 3", vehicle_registration="B1003AA", capacity_label="16KL", vehicle_type_tag=16, number_of_compartments=2, depot_id="D1", active_status="ACTIVE"),
            MasterMT(mt_id="INACTIVE", vehicle_name_raw="Inactive", vehicle_registration="B1999AA", vehicle_type_tag=8, number_of_compartments=1, depot_id="D1", active_status="INACTIVE"),
            MasterMT(mt_id="OTHER-MT", vehicle_name_raw="Other", vehicle_registration="B2001BB", vehicle_type_tag=8, number_of_compartments=1, depot_id="D2", active_status="ACTIVE"),
            MasterProduct(product_id="P-PERTAMAX", product_name="PERTAMAX", normalized_product="PERTAMAX", active_status="ACTIVE"),
            MasterSPBU(spbu_id="A", spbu_code="SPBU-A", spbu_name="A", latitude=-6.21, longitude=106.85, master_travel_time_min=10, vehicle_type_tag=8, primary_depot_id="D1"),
            MasterSPBU(spbu_id="B", spbu_code="SPBU-B", spbu_name="B", latitude=-6.22, longitude=106.86, master_travel_time_min=10, vehicle_type_tag=8, primary_depot_id="D1"),
            MasterSPBU(spbu_id="C", spbu_code="SPBU-C", spbu_name="C", latitude=-6.23, longitude=106.87, master_travel_time_min=10, vehicle_type_tag=8, primary_depot_id="D1"),
            MasterSPBU(spbu_id="LIMITED", spbu_code="SPBU-LIMITED", spbu_name="Limited", latitude=-6.24, longitude=106.88, master_travel_time_min=10, vehicle_type_tag=16, primary_depot_id="D1"),
            MasterSPBU(spbu_id="OTHER", spbu_code="SPBU-OTHER", spbu_name="Other", vehicle_type_tag=8, primary_depot_id="D2"),
        ]
    )
    model = MLBehavioralModel(
        model_id="M1",
        model_name="Ready Model",
        model_version=1,
        depot_id="D1",
        training_start_date=date(2026, 1, 1),
        training_end_date=date(2026, 6, 30),
        training_shipment_count=100,
        training_spbu_count=4,
        cluster_count=3,
        average_membership_probability=0.9,
        feature_weights={"tag": 0.4, "shift": 0.25, "pairing": 0.35},
        shift_definition_snapshot=[
            {"shift_id": "s1", "name": "Shift 1", "start_time": "00:00", "end_time": "05:59"},
            {"shift_id": "s2", "name": "Shift 2", "start_time": "06:00", "end_time": "11:59"},
            {"shift_id": "s3", "name": "Shift 3", "start_time": "12:00", "end_time": "17:59"},
            {"shift_id": "s4", "name": "Shift 4", "start_time": "18:00", "end_time": "23:59"},
        ],
        model_status="ACTIVE",
    )
    session.add(model)
    session.commit()
    for spbu_id, cluster, shift in (("A", 0, "Shift 1"), ("B", 0, "Shift 1"), ("C", 1, "Shift 2"), ("LIMITED", 1, "Shift 2")):
        session.add(
            MLSPBUClusterAssignment(
                assignment_id=f"AS-{spbu_id}", model_id="M1", depot_id="D1", spbu_id=spbu_id,
                cluster_id=cluster, cluster_label=f"Cluster {cluster + 1}", membership_probability=0.9,
                is_noise=False, dominant_shift=shift,
            )
        )
    session.commit()
    return model


def run_prediction(session, lo_rows: list[list], mt_rows: list[list], parameters: dict | None = None) -> dict:
    normalized_lo_rows = [
        row if len(row) >= 5 else [*row[:3], "PERTAMAX", row[3] if len(row) == 4 else 8]
        for row in lo_rows
    ]
    return create_prediction_run(
        session,
        depot_id="D1",
        model_id="M1",
        loading_order_content=workbook(["loading_order_no", "shipment_start_datetime", "spbu_no", "product", "order_quantity_kl"], normalized_lo_rows),
        loading_order_filename="lo.xlsx",
        availability_content=workbook(["vehicle_registration_no", "initial_available_datetime"], mt_rows),
        availability_filename="mt.xlsx",
        parameters=parameters,
        created_by="tester",
    )


def test_prediction_run_can_be_queued_then_processed() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        queued = enqueue_prediction_run(
            session,
            depot_id="D1",
            model_id="M1",
            loading_order_content=workbook(
                ["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"],
                [["LO1", "2026-08-22 09:00:00", "SPBU-A", 8]],
            ),
            loading_order_filename="lo.xlsx",
            availability_content=workbook(
                ["vehicle_registration_no", "initial_available_datetime"],
                [["B1001AA", "2026-08-22 08:00:00"]],
            ),
            availability_filename="mt.xlsx",
            parameters=None,
            created_by="tester",
        )
        assert queued["status"] == "QUEUED"
        assert get_prediction_run_status(session, queued["id"])["status"] == "QUEUED"

        claimed = claim_next_prediction_job(session, worker_id="test-worker", lease_seconds=30)
        assert claimed is not None
        assert claimed.run_id == queued["id"]
        assert heartbeat_prediction_job(
            session,
            run_id=claimed.run_id,
            lease_token=claimed.lease_token,
            lease_seconds=30,
        )

        completed = process_prediction_run(session, queued["id"], lease_token=claimed.lease_token)
        assert completed["status"] == "COMPLETED"
        assert completed["summary"]["loading_orders"] == 1
        assert get_prediction_run_status(session, queued["id"])["status"] == "COMPLETED"
        assert session.get(PredictionJob, queued["id"]).status == "COMPLETED"


def test_multiple_predictions_can_queue_without_waiting_for_active_job() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        queued_runs = [
            enqueue_prediction_run(
                session,
                depot_id="D1",
                model_id="M1",
                loading_order_content=workbook(
                    ["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"],
                    [[f"LO{index}", "2026-08-22 09:00:00", "SPBU-A", 8]],
                ),
                loading_order_filename=f"lo-{index}.xlsx",
                availability_content=workbook(
                    ["vehicle_registration_no", "initial_available_datetime"],
                    [["B1001AA", "2026-08-22 08:00:00"]],
                ),
                availability_filename=f"mt-{index}.xlsx",
                parameters=None,
                created_by="tester",
            )
            for index in (1, 2)
        ]
        assert [get_prediction_run_status(session, row["id"])["status"] for row in queued_runs] == ["QUEUED", "QUEUED"]
        first = claim_next_prediction_job(session, worker_id="worker-1", lease_seconds=30)
        assert first is not None
        assert get_prediction_run_status(session, first.run_id)["status"] == "RUNNING"
        queued_id = next(row["id"] for row in queued_runs if row["id"] != first.run_id)
        assert get_prediction_run_status(session, queued_id)["status"] == "QUEUED"


def test_prediction_result_is_paginated_and_candidates_are_lazy() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        result = run_prediction(
            session,
            [
                ["LO1", "2026-08-22 05:30:00", "SPBU-A"],
                ["LO2", "2026-08-22 10:00:00", "SPBU-C"],
            ],
            [["B1001AA", "2026-08-22 05:00:00"], ["B1002AA", "2026-08-22 05:00:00"]],
        )
        page = get_prediction_run(
            session,
            result["id"],
            shipment_page=1,
            shipment_page_size=1,
            include_candidates=False,
        )
        assert page["shipment_pagination"] == {
            "page": 1,
            "page_size": 1,
            "total": 2,
            "total_pages": 2,
            "shift_id": None,
        }
        assert len(page["shipments"]) == 1
        assert page["shipments"][0]["candidates"] == []
        assert page["shipments"][0]["candidates_loaded"] is False
        line = page["shipments"][0]["lines"][0]
        assert line["spbu_no"] == "SPBU-A"
        assert line["product_id"] == "P-PERTAMAX"
        assert line["product_name"] == "PERTAMAX"
        assert line["cluster_id"] == 0
        assert line["cluster_number"] == 1
        assert line["cluster_label"] == "Cluster 1"
        shipment_option = next(
            item for item in page["shipment_options"]
            if item["id"] == page["shipments"][0]["id"]
        )
        assert shipment_option["spbus"] == [
            {
                "spbu_id": "A",
                "spbu_no": "SPBU-A",
                "cluster_id": 0,
                "cluster_number": 1,
                "cluster_label": "Cluster 1",
            }
        ]
        assert all("route_geometry" not in trip for trip in page["trips"])
        candidates = get_prediction_shipment_candidates(session, result["id"], page["shipments"][0]["id"])
        assert candidates["candidates"]


def test_prediction_api_persists_queue_task_with_202() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)

    def override_db():
        with Session() as session:
            yield session

    app.dependency_overrides[get_db] = override_db
    try:
        client = TestClient(app)
        response = client.post(
            "/api/v1/phase6/predictions",
            data={"depot_id": "D1", "model_id": "M1", "parameters": "{}"},
            files={
                "loading_order_file": (
                    "lo.xlsx",
                    workbook(
                        ["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"],
                        [["LO1", "2026-08-22 09:00:00", "SPBU-A", 8]],
                    ),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                "mt_availability_file": (
                    "mt.xlsx",
                    workbook(
                        ["vehicle_registration_no", "initial_available_datetime"],
                        [["B1001AA", "2026-08-22 08:00:00"]],
                    ),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        assert response.status_code == 202
        queued = response.json()
        assert queued["status"] == "QUEUED"
        with Session() as session:
            status_payload = get_prediction_run_status(session, queued["id"])
            assert status_payload["status"] == "QUEUED"
            assert status_payload["queue"]["attempt_count"] == 0
    finally:
        app.dependency_overrides.clear()


def test_managed_loading_order_workbook_api_returns_excel_file() -> None:
    response = TestClient(app).post(
        "/api/v1/phase6/loading-orders/workbook",
        json={
            "rows": [
                {
                    "loading_order_no": "LO-API-1",
                    "shipment_start_datetime": "2026-08-22 09:00:00",
                    "spbu_no": "SPBU-A",
                    "product": "PERTAMAX",
                    "order_quantity_kl": 8,
                }
            ]
        },
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    sheet = load_workbook(BytesIO(response.content), read_only=True, data_only=True).active
    assert list(sheet.iter_rows(min_row=2, values_only=True)) == [("LO-API-1", "2026-08-22 09:00:00", "SPBU-A", "PERTAMAX", 8)]


def test_managed_mt_availability_workbook_api_returns_excel_file() -> None:
    response = TestClient(app).post(
        "/api/v1/phase6/mt-availability/workbook",
        json={
            "rows": [
                {
                    "vehicle_registration_no": "B1001AA",
                    "initial_available_datetime": "2026-08-22 00:00:00",
                }
            ]
        },
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    sheet = load_workbook(BytesIO(response.content), read_only=True, data_only=True).active
    assert list(sheet.iter_rows(min_row=2, values_only=True)) == [("B1001AA", "2026-08-22 00:00:00")]


def test_master_spbu_options_filter_selected_depot_and_active_status() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        session.add(
            MasterSPBU(
                spbu_id="INACTIVE-SPBU",
                spbu_code="SPBU-INACTIVE",
                spbu_name="Inactive SPBU",
                primary_depot_id="D1",
                active_status="INACTIVE",
            )
        )
        session.commit()

    def override_db():
        with Session() as session:
            yield session

    app.dependency_overrides[get_db] = override_db
    try:
        response = TestClient(app).get("/api/v1/master/spbu?depot_id=D1&active_only=true&limit=10000")
        assert response.status_code == 200
        codes = {row["spbu_code"] for row in response.json()}
        assert {"SPBU-A", "SPBU-B", "SPBU-C", "SPBU-LIMITED"} <= codes
        assert "SPBU-OTHER" not in codes
        assert "SPBU-INACTIVE" not in codes
    finally:
        app.dependency_overrides.clear()


def test_master_mt_options_filter_selected_depot_and_active_status() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        session.add(
            MasterMT(
                mt_id="INVALID-CAPACITY",
                vehicle_name_raw="Invalid Capacity",
                vehicle_registration="B1004AA",
                vehicle_type_tag=10,
                depot_id="D1",
                active_status="ACTIVE",
            )
        )
        session.commit()

    def override_db():
        with Session() as session:
            yield session

    app.dependency_overrides[get_db] = override_db
    try:
        client = TestClient(app)
        response = client.get("/api/v1/master/mt?depot_id=D1&active_only=true&limit=10000")
        assert response.status_code == 200
        registrations = {row["vehicle_registration"] for row in response.json()}
        assert registrations == {"B1001AA", "B1002AA", "B1003AA", "B1004AA"}

        availability_response = client.get("/api/v1/phase6/master-mt-availability?depot_id=D1")
        assert availability_response.status_code == 200
        availability = {row["vehicle_registration"]: row for row in availability_response.json()}
        assert set(availability) == {"B1001AA", "B1002AA", "B1003AA", "B1004AA"}
        assert all(availability[registration]["phase6_eligible"] is True for registration in {"B1001AA", "B1002AA", "B1003AA"})
        assert availability["B1004AA"]["phase6_eligible"] is False
        assert "MT_CAPACITY_NOT_8_KL_MULTIPLE" in availability["B1004AA"]["phase6_failed_rules"]

        models_response = client.get("/api/v1/phase6/models?depot_id=D1")
        assert models_response.status_code == 200
        assert models_response.json()[0]["shift_definition_snapshot"][0]["start_time"] == "00:00"
    finally:
        app.dependency_overrides.clear()


def test_stale_worker_job_is_requeued_then_fails_at_retry_limit() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        queued = enqueue_prediction_run(
            session,
            depot_id="D1",
            model_id="M1",
            loading_order_content=workbook(
                ["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"],
                [["LO1", "2026-08-22 09:00:00", "SPBU-A", 8]],
            ),
            loading_order_filename="lo.xlsx",
            availability_content=workbook(
                ["vehicle_registration_no", "initial_available_datetime"],
                [["B1001AA", "2026-08-22 08:00:00"]],
            ),
            availability_filename="mt.xlsx",
            parameters=None,
            created_by="tester",
        )
        job = session.get(PredictionJob, queued["id"])
        job.max_attempts = 2
        session.commit()

        first = claim_next_prediction_job(session, worker_id="worker-1", lease_seconds=1)
        assert first is not None
        recovered = recover_stale_prediction_jobs(session, now=utc_now() + timedelta(seconds=2))
        assert recovered == {"requeued": 1, "failed": 0}
        assert get_prediction_run_status(session, queued["id"])["status"] == "QUEUED"

        second = claim_next_prediction_job(session, worker_id="worker-2", lease_seconds=1)
        assert second is not None
        recovered = recover_stale_prediction_jobs(session, now=utc_now() + timedelta(seconds=2))
        assert recovered == {"requeued": 0, "failed": 1}
        payload = get_prediction_run_status(session, queued["id"])
        assert payload["status"] == "FAILED"
        assert payload["error_code"] == "WORKER_HEARTBEAT_TIMEOUT"


def test_timestamp_validation_derives_shift_and_rejects_bad_inputs() -> None:
    Session = make_session()
    with Session() as session:
        model = seed(session)
        result = validate_loading_orders(
            session, depot_id="D1", model=model,
            content=workbook(
                ["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"],
                [["LO1", "2026-08-22 05:30:00", "SPBU-A", 8], ["LO1", "bad", "UNKNOWN", 8], ["LO3", "2026-08-22 07:00:00", "SPBU-OTHER", 8]],
            ),
            file_name="lo.xlsx",
        )
        codes = {issue["error_code"] for issue in result["issues"]}
        assert {"DUPLICATE_LOADING_ORDER", "INVALID_DATETIME", "SPBU_NOT_FOUND", "SPBU_DEPOT_MISMATCH"} <= codes

        valid = validate_loading_orders(
            session, depot_id="D1", model=model,
            content=workbook(["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"], [["LO4", "2026-08-22 07:00:00", "SPBU-A", 8]]),
            file_name="lo.xlsx",
        )
        assert valid["status"] == "PASS"
        assert valid["normalized_rows"][0]["shift_id"] == "s2"
        assert valid["normalized_rows"][0]["shipment_start_datetime"].endswith("+00:00")


def test_loading_order_quantity_must_be_present_and_exactly_8_kl() -> None:
    Session = make_session()
    with Session() as session:
        model = seed(session)
        result = validate_loading_orders(
            session,
            depot_id="D1",
            model=model,
            content=workbook(
                ["loading_order_no", "shipment_start_datetime", "spbu_no", "order_quantity_kl"],
                [
                    ["LO1", "2026-08-22 07:00:00", "SPBU-A", 10],
                    ["LO2", "2026-08-22 07:00:00", "SPBU-A", None],
                    ["LO3", "2026-08-22 07:00:00", "SPBU-A", 16],
                ],
            ),
            file_name="lo.xlsx",
        )
        codes = {issue["error_code"] for issue in result["issues"]}
        assert "ORDER_QUANTITY_MUST_BE_8_KL" in codes
        assert "REQUIRED_VALUE_EMPTY" in codes
        assert result["blocking_error_count"] > 0


def test_loading_order_product_maps_to_canonical_master_and_rejects_unknown() -> None:
    Session = make_session()
    with Session() as session:
        model = seed(session)
        result = validate_loading_orders(
            session,
            depot_id="D1",
            model=model,
            content=workbook(
                ["loading_order_no", "shipment_start_datetime", "spbu_no", "product", "order_quantity_kl"],
                [
                    ["LO-PRODUCT-1", "2026-08-22 07:00:00", "SPBU-A", "PERTAMAX", 8],
                    ["LO-PRODUCT-2", "2026-08-22 07:15:00", "SPBU-B", "UNKNOWN PRODUCT", 8],
                ],
            ),
            file_name="lo-product.xlsx",
        )
        assert result["normalized_rows"][0]["product_id"] == "P-PERTAMAX"
        assert result["normalized_rows"][0]["product_name"] == "PERTAMAX"
        assert "PRODUCT_NOT_FOUND" in {issue["error_code"] for issue in result["issues"]}


def test_managed_loading_order_workbook_round_trips_editable_rows() -> None:
    Session = make_session()
    with Session() as session:
        model = seed(session)
        content = loading_order_workbook(
            [
                {
                    "loading_order_no": "LO-MANAGED-1",
                    "shipment_start_datetime": "2026-08-22 07:00:00",
                    "spbu_no": "SPBU-A",
                    "product": "PERTAMAX",
                    "order_quantity_kl": 8,
                },
                {
                    "loading_order_no": "LO-MANAGED-2",
                    "shipment_start_datetime": "2026-08-22 07:15:00",
                    "spbu_no": "SPBU-B",
                    "product": "PERTAMAX",
                    "order_quantity_kl": 8,
                },
            ]
        )
        result = validate_loading_orders(
            session,
            depot_id="D1",
            model=model,
            content=content,
            file_name="phase6-managed-loading-orders.xlsx",
        )
        assert result["status"] == "PASS"
        assert [row["loading_order_no"] for row in result["editable_rows"]] == ["LO-MANAGED-1", "LO-MANAGED-2"]
        assert [row["product_name"] for row in result["normalized_rows"]] == ["PERTAMAX", "PERTAMAX"]
        assert [row["order_quantity_kl"] for row in result["normalized_rows"]] == [8, 8]


def test_mt_timestamp_validation_duplicate_inactive_and_depot_errors() -> None:
    Session = make_session()
    with Session() as session:
        model = seed(session)
        result = validate_mt_availability(
            session, depot_id="D1", model=model,
            content=workbook(
                ["vehicle_registration_no", "initial_available_datetime"],
                [["B1001AA", "2026-08-22 05:00:00"], ["B1001AA", "bad"], ["B1999AA", "2026-08-22 05:00:00"], ["B2001BB", "2026-08-22 05:00:00"]],
            ),
            file_name="mt.xlsx",
        )
        codes = {issue["error_code"] for issue in result["issues"]}
        assert {"DUPLICATE_VEHICLE_AVAILABILITY", "INVALID_AVAILABLE_DATETIME", "VEHICLE_INACTIVE", "VEHICLE_DEPOT_MISMATCH"} <= codes


def test_strict_start_multi_trip_reuses_vehicle_without_overlap() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        result = run_prediction(
            session,
            [["LO1", "2026-08-22 05:30:00", "SPBU-A"], ["LO2", "2026-08-22 10:00:00", "SPBU-C"]],
            [["B1001AA", "2026-08-22 05:00:00"]],
            {"assignment_mode": "STRICT_START"},
        )
        assert [trip["vehicle_id"] for trip in result["trips"]] == ["T1", "T1"]
        assert [trip["trip_number"] for trip in result["trips"]] == [1, 2]
        first, second = result["trips"]
        assert datetime.fromisoformat(first["next_available_datetime"]) <= datetime.fromisoformat(second["predicted_departure_datetime"])
        assert result["summary"]["multi_trip_mt"] == 1
        assert sum(row["delivered_kl"] for row in result["hourly_distribution"]) == 16
        assert result["hourly_distribution"][-1]["cumulative_kl"] == 16
        assert len(result["geographic_routes"]["routes"]) == 2
        assert result["geographic_routes"]["sequence_policy"] == "NEAREST_TO_FARTHEST_FROM_DEPOT"
        assert all(route["vehicle_id"] == "T1" for route in result["geographic_routes"]["routes"])
        assert all(route["route_geometry_source"] == "MASTER_COORDINATE_FALLBACK" for route in result["geographic_routes"]["routes"])
        assert all(route["uses_road_geometry"] is False for route in result["geographic_routes"]["routes"])


def test_strict_start_rejects_not_yet_available_vehicle() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        result = run_prediction(
            session,
            [["LO1", "2026-08-22 09:00:00", "SPBU-A"]],
            [["B1001AA", "2026-08-22 09:30:00"]],
            {"assignment_mode": "STRICT_START"},
        )
        assert result["trips"][0]["assignment_status"] == "UNASSIGNED"
        assert result["trips"][0]["unassigned_reason"] == "NO_MT_AVAILABLE_AT_REQUIRED_TIME"


def test_allow_delay_assigns_with_expected_departure() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        result = run_prediction(
            session,
            [["LO1", "2026-08-22 09:00:00", "SPBU-A"]],
            [["B1001AA", "2026-08-22 09:20:00"]],
            {"assignment_mode": "ALLOW_DELAY", "maximum_allowed_delay_minutes": 30},
        )
        trip = result["trips"][0]
        assert trip["assignment_status"] == "ASSIGNED_WITH_DELAY"
        assert trip["delay_minutes"] == 20
        assert datetime.fromisoformat(trip["predicted_departure_datetime"]).astimezone(timezone.utc).hour == 2


def test_multi_spbu_master_compatibility_is_intersection() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        # C and LIMITED are deliberately paired by allowing a zero confidence threshold.
        result = run_prediction(
            session,
            [["LO1", "2026-08-22 09:00:00", "SPBU-C"], ["LO2", "2026-08-22 09:00:00", "SPBU-LIMITED"]],
            [["B1003AA", "2026-08-22 08:00:00"]],
            {"minimum_prediction_confidence": 0},
        )
        assert result["shipments"][0]["candidates"][0]["compatibility_status"] == "FAIL"
        assert result["trips"][0]["unassigned_reason"] == "NO_COMPATIBLE_MT"


def test_capacity_aware_grouping_builds_three_lo_shipment_for_24_kl_mt() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        session.add(
            MasterMT(
                mt_id="T24",
                vehicle_name_raw="Truck 24",
                vehicle_registration="B1024AA",
                capacity_label="24KL",
                vehicle_type_tag=24,
                number_of_compartments=3,
                depot_id="D1",
                active_status="ACTIVE",
            )
        )
        for suffix, latitude in (("D", -6.25), ("E", -6.26), ("F", -6.27)):
            session.add(
                MasterSPBU(
                    spbu_id=suffix,
                    spbu_code=f"SPBU-{suffix}",
                    spbu_name=suffix,
                    latitude=latitude,
                    longitude=106.88,
                    master_travel_time_min=10,
                    vehicle_type_tag=24,
                    primary_depot_id="D1",
                )
            )
            session.flush()
            session.add(
                MLSPBUClusterAssignment(
                    assignment_id=f"AS-{suffix}",
                    model_id="M1",
                    depot_id="D1",
                    spbu_id=suffix,
                    cluster_id=2,
                    cluster_label="Cluster 3",
                    membership_probability=0.9,
                    is_noise=False,
                    dominant_shift="Shift 2",
                )
            )
        session.commit()

        result = run_prediction(
            session,
            [
                ["LO-D", "2026-08-22 09:00:00", "SPBU-D", 8],
                ["LO-E", "2026-08-22 09:05:00", "SPBU-E", 8],
                ["LO-F", "2026-08-22 09:10:00", "SPBU-F", 8],
            ],
            [["B1024AA", "2026-08-22 08:00:00"]],
        )
        assert len(result["shipments"]) == 1
        shipment = result["shipments"][0]
        assert len(shipment["lines"]) == 3
        assert shipment["total_order_kl"] == 24
        assert shipment["required_compartments"] == 3
        assert shipment["explanation"]["grouping_method"] == "CAPACITY_TIME_ROUTE_SET_PACKING"
        assert shipment["explanation"]["optimizer_method"] in {
            "SCIPY_MILP_SET_PACKING",
            "DETERMINISTIC_GREEDY_SET_PACKING_FALLBACK",
        }
        assert shipment["explanation"]["group_optimization"]["route_feasible"] is True
        assert shipment["explanation"]["group_optimization"]["time_span_minutes"] == 10
        assert shipment["assignment"]["assigned_vehicle_id"] == "T24"
        assert shipment["explanation"]["capacity_iteration"]["tier_capacity_kl"] == 24
        assert result["model"]["capacity_iteration_summary"][1]["assigned_shipments"] == 1
        assert shipment["trip"]["estimated_visit_sequence"] == ["SPBU-D", "SPBU-E", "SPBU-F"]
        route = result["geographic_routes"]["routes"][0]
        assert [stop["spbu_code"] for stop in route["stops"]] == ["SPBU-D", "SPBU-E", "SPBU-F"]
        assert route["points"][0]["type"] == "DEPOT"
        assert route["points"][-1]["type"] == "DEPOT_RETURN"


def test_iterative_capacity_assignment_regroups_unassigned_32_kl_into_tag_compatible_16_kl() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        session.add(
            MasterMT(
                mt_id="T4",
                vehicle_name_raw="Truck 16 without required project tag",
                vehicle_registration="B1004AA",
                capacity_label="16KL",
                vehicle_type_tag=16,
                number_of_compartments=2,
                depot_id="D1",
                active_status="ACTIVE",
            )
        )
        session.add(MasterTagType(tag_type_id="PROJECT", code="PROJECT", name="Project"))
        session.flush()
        session.add(MasterTag(tag_id="TAG-REQ", tag_type_id="PROJECT", tag_value="Required", normalized_tag="REQUIRED"))
        session.flush()
        session.add_all(
            [
                BridgeSPBUTag(spbu_id="LIMITED", tag_id="TAG-REQ"),
                BridgeMTTag(mt_id="T3", tag_id="TAG-REQ"),
            ]
        )
        session.commit()

        result = run_prediction(
            session,
            [
                ["LO1", "2026-08-22 09:00:00", "SPBU-LIMITED", 8],
                ["LO2", "2026-08-22 09:01:00", "SPBU-LIMITED", 8],
                ["LO3", "2026-08-22 09:02:00", "SPBU-LIMITED", 8],
                ["LO4", "2026-08-22 09:03:00", "SPBU-LIMITED", 8],
            ],
            [["B1003AA", "2026-08-22 08:00:00"], ["B1004AA", "2026-08-22 08:00:00"]],
            {"assignment_mode": "ALLOW_DELAY", "maximum_allowed_delay_minutes": 300},
        )

        assert [shipment["total_order_kl"] for shipment in result["shipments"]] == [16, 16]
        assert all(shipment["assignment"]["assigned_vehicle_id"] == "T3" for shipment in result["shipments"])
        assert {trip["assignment_status"] for trip in result["trips"]} == {"ASSIGNED", "ASSIGNED_WITH_DELAY"}
        assert result["summary"]["assigned_loading_orders"] == 4
        assert result["summary"]["assigned_order_kl"] == 32
        assert result["model"]["capacity_iteration_order"] == [32, 24, 16, 8]
        iterations = {row["tier_capacity_kl"]: row for row in result["model"]["capacity_iteration_summary"]}
        assert iterations[32]["assigned_shipments"] == 0
        assert iterations[16]["assigned_shipments"] == 2
        assert iterations[16]["carried_forward_loading_orders"] == 0
        assert all(shipment["explanation"]["capacity_iteration"]["spbu_cluster_evidence_required_for_multi_lo"] for shipment in result["shipments"])
        assert all(shipment["explanation"]["capacity_iteration"]["spbu_mt_master_tag_compatibility_required"] for shipment in result["shipments"])
        failed_untagged = [
            candidate
            for shipment in result["shipments"]
            for candidate in shipment["candidates"]
            if candidate["vehicle_id"] == "T4"
        ]
        assert failed_untagged
        assert all(candidate["compatibility_status"] == "FAIL" for candidate in failed_untagged)
        assert all("PROJECT_TAGS" in candidate["explanation"]["failed_rules"] for candidate in failed_untagged)


def test_one_8_kl_lo_cannot_use_16_kl_mt_when_full_load_is_required() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        result = run_prediction(
            session,
            [["LO1", "2026-08-22 09:00:00", "SPBU-LIMITED", 8]],
            [["B1003AA", "2026-08-22 08:00:00"]],
        )
        candidate = result["shipments"][0]["candidates"][0]
        assert candidate["compatibility_status"] == "FAIL"
        assert candidate["exclusion_reason"] == "CAPACITY_COMPARTMENT_MISMATCH"
        assert candidate["explanation"]["capacity_policy"] == "EXACT_COMPARTMENT_MATCH"
        assert candidate["explanation"]["shipment_required_compartments"] == 1
        assert candidate["explanation"]["mt_number_of_compartments"] == 2
        assert result["shipments"][0]["explanation"]["capacity_iteration"]["tier_capacity_kl"] == 8
        assert result["trips"][0]["vehicle_id"] is None
        assert result["trips"][0]["unassigned_reason"] == "NO_COMPATIBLE_MT"
        assert result["summary"]["assigned_loading_orders"] == 0
        assert result["summary"]["assigned_order_kl"] == 0


def test_demo_loading_order_uses_timestamps_and_requested_total() -> None:
    Session = make_session()
    with Session() as session:
        model = seed(session)
        session.add(
            MasterSPBU(
                spbu_id="UNCOVERED",
                spbu_code="SPBU-UNCOVERED",
                spbu_name="Uncovered",
                latitude=-6.30,
                longitude=106.90,
                vehicle_type_tag=8,
                primary_depot_id="D1",
                active_status="ACTIVE",
            )
        )
        session.add(
            MasterSPBU(
                spbu_id="COLD",
                spbu_code="SPBU-COLD",
                spbu_name="Cold Start",
                latitude=-6.31,
                longitude=106.91,
                vehicle_type_tag=8,
                primary_depot_id="D1",
                active_status="ACTIVE",
            )
        )
        session.flush()
        session.add(
            MLSPBUClusterAssignment(
                assignment_id="AS-COLD",
                model_id="M1",
                depot_id="D1",
                spbu_id="COLD",
                cluster_id=0,
                cluster_label="Cluster 1",
                membership_probability=0.99,
                is_noise=False,
                dominant_shift="Shift 1",
                shipment_observation_count=0,
                coverage_source="ACTIVE_MASTER_COLD_START",
                history_eligible=False,
            )
        )
        session.commit()
        with pytest.raises(HTTPException) as raised:
            generate_demo_loading_orders(session, depot_id="D1", model=model, total_order_kl=18, random_seed=42)
        assert raised.value.detail["code"] == "DEMO_TOTAL_ORDER_NOT_8_KL_MULTIPLE"

        content, filename = generate_demo_loading_orders(session, depot_id="D1", model=model, total_order_kl=24, random_seed=42)
        sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
        assert [cell.value for cell in next(sheet.iter_rows(max_row=1))] == ["loading_order_no", "shipment_start_datetime", "spbu_no", "spbu_name", "product", "order_quantity_kl"]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert filename.startswith("phase6-demo-loading-order-")
        assert [row[4] for row in rows] == ["PERTAMAX", "PERTAMAX", "PERTAMAX"]
        assert [row[5] for row in rows] == [8, 8, 8]
        assert {row[2] for row in rows} <= {"SPBU-A", "SPBU-B", "SPBU-C", "SPBU-LIMITED"}
        assert "SPBU-COLD" not in {row[2] for row in rows}
        generated_times = [datetime.fromisoformat(row[1]) for row in rows]
        assert (max(generated_times) - min(generated_times)).total_seconds() <= 2 * 60
        validated = validate_loading_orders(session, depot_id="D1", model=model, content=content, file_name=filename)
        assert validated["status"] == "PASS"
        assert {row["product_name"] for row in validated["normalized_rows"]} == {"PERTAMAX"}
        assert sum(row["order_quantity_kl"] for row in validated["normalized_rows"]) == 24


def test_demo_mt_availability_selects_random_active_mt_near_capacity_target() -> None:
    Session = make_session()
    with Session() as session:
        model = seed(session)
        content, filename = generate_demo_mt_availability(session, depot_id="D1", model=model, total_capacity_kl=24, random_seed=42)
        sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
        assert [cell.value for cell in next(sheet.iter_rows(max_row=1))][:3] == [
            "vehicle_registration_no",
            "initial_available_datetime",
            "capacity_kl",
        ]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert filename.startswith("phase6-demo-mt-availability-")
        assert len(rows) == 2
        assert sum(row[2] for row in rows) == 24
        assert {row[0] for row in rows} <= {"B1001AA", "B1002AA", "B1003AA"}
        availability_times = [datetime.fromisoformat(row[1]) for row in rows]
        assert len(set(availability_times)) == 1
        assert availability_times[0].hour == 0
        assert availability_times[0].minute == 0
        validated = validate_mt_availability(session, depot_id="D1", model=model, content=content, file_name=filename)
        assert validated["status"] == "PASS"
        assert sum(row["capacity_kl"] for row in validated["normalized_rows"]) == 24

        random_content, _ = generate_demo_mt_availability(
            session,
            depot_id="D1",
            model=model,
            total_capacity_kl=24,
            random_availability=True,
            random_seed=42,
        )
        random_sheet = load_workbook(BytesIO(random_content), read_only=True, data_only=True).active
        random_times = [datetime.fromisoformat(row[1]) for row in random_sheet.iter_rows(min_row=2, values_only=True)]
        assert len(set(random_times)) > 1
        depot_close = availability_times[0].replace(hour=23, minute=59)
        assert all(availability_times[0] <= value <= depot_close for value in random_times)

        near_content, _ = generate_demo_mt_availability(session, depot_id="D1", model=model, total_capacity_kl=18, random_seed=7)
        near_sheet = load_workbook(BytesIO(near_content), read_only=True, data_only=True).active
        assert sum(row[2] for row in near_sheet.iter_rows(min_row=2, values_only=True)) == 16


def test_google_client_is_drive_only_and_rejects_truck() -> None:
    captured = {}

    def handler(request: httpx.Request) -> httpx.Response:
        captured.update(json.loads(request.content))
        assert request.headers["x-goog-api-key"] == "secret-test-key"
        return httpx.Response(200, json={"routes": [{
            "distanceMeters": 1000,
            "duration": "600s",
            "staticDuration": "550s",
            "polyline": {"geoJsonLinestring": {"type": "LineString", "coordinates": [[106.8, -6.2], [106.85, -6.25], [106.9, -6.3]]}},
        }]})

    client = GoogleRoutesClient("secret-test-key", transport=httpx.MockTransport(handler))
    response = client.compute_route(
        origin=(-6.2, 106.8), destination=(-6.3, 106.9), departure_datetime=datetime(2099, 1, 1, tzinfo=timezone.utc),
        routing_mode="DRIVE", routing_preference="TRAFFIC_AWARE",
    )
    assert captured["travelMode"] == "DRIVE"
    assert captured["routingPreference"] == "TRAFFIC_AWARE"
    assert captured["polylineQuality"] == "OVERVIEW"
    assert captured["polylineEncoding"] == "GEO_JSON_LINESTRING"
    assert "routeModifiers" not in captured
    assert response["duration_seconds"] == 600
    assert response["route_geometry"] == [
        {"latitude": -6.2, "longitude": 106.8},
        {"latitude": -6.25, "longitude": 106.85},
        {"latitude": -6.3, "longitude": 106.9},
    ]
    assert response["route_geometry_source"] == "GOOGLE_ROUTES_GEOJSON"

    with pytest.raises(GoogleRoutesError) as raised:
        client.compute_route(
            origin=(-6.2, 106.8), destination=(-6.3, 106.9), departure_datetime=None,
            routing_mode="TRUCK", routing_preference="TRAFFIC_AWARE",
        )
    assert raised.value.code == "UNSUPPORTED_ROUTING_MODE"


class FakeRoutesClient:
    def __init__(self):
        self.calls = 0
        self.routing_modes: list[str] = []

    def compute_route(self, **kwargs):
        self.calls += 1
        self.routing_modes.append(kwargs["routing_mode"])
        return {
            "distance_meters": 1000,
            "duration_seconds": 600,
            "static_duration_seconds": 550,
            "route_geometry": [
                {"latitude": kwargs["origin"][0], "longitude": kwargs["origin"][1]},
                {"latitude": kwargs["destination"][0], "longitude": kwargs["destination"][1]},
            ],
            "route_geometry_source": "GOOGLE_ROUTES_GEOJSON",
            "restrictions_partially_ignored": False,
            "warnings": [],
        }


def test_route_cache_uses_generic_drive_profile_across_vehicles() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        mt1, mt2 = session.get(MasterMT, "T1"), session.get(MasterMT, "T2")
        configuration = get_google_routes_configuration(session)
        configuration.routing_mode = "DRIVE"
        fake = FakeRoutesClient()
        metrics = {}
        service = Phase6RouteEstimationService(session, configuration=configuration, model_id="M1", metrics=metrics, google_client=fake)
        depot, spbu = session.get(MasterDepot, "D1"), session.get(MasterSPBU, "A")
        departure = datetime(2099, 1, 1, 1, tzinfo=timezone.utc)
        service.estimate_trip(depot=depot, spbus=[spbu], mt=mt1, predicted_departure_datetime=departure, max_exact_sequence_stops=4)
        session.flush()
        service.estimate_trip(depot=depot, spbus=[spbu], mt=mt1, predicted_departure_datetime=departure, max_exact_sequence_stops=4)
        service.estimate_trip(depot=depot, spbus=[spbu], mt=mt2, predicted_departure_datetime=departure, max_exact_sequence_stops=4)
        assert fake.calls == 2  # outbound + return once; DRIVE cache is vehicle agnostic
        assert metrics["google_routes_cache_hit_count"] == 4


def test_route_cache_reuses_pending_rows_before_session_flush() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        configuration = get_google_routes_configuration(session)
        fake = FakeRoutesClient()
        service = Phase6RouteEstimationService(
            session,
            configuration=configuration,
            model_id="M1",
            google_client=fake,
        )
        arguments = {
            "depot": session.get(MasterDepot, "D1"),
            "spbus": [session.get(MasterSPBU, "A")],
            "mt": session.get(MasterMT, "T1"),
            "predicted_departure_datetime": datetime(2099, 1, 1, 1, tzinfo=timezone.utc),
            "max_exact_sequence_stops": 4,
        }
        service.estimate_trip(**arguments)
        service.estimate_trip(**arguments)
        session.flush()
        assert fake.calls == 2


def test_route_estimation_forces_drive_even_with_legacy_truck_configuration() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        configuration = get_google_routes_configuration(session)
        configuration.routing_mode = "TRUCK"
        configuration.truck_routing_status = "AVAILABLE"
        fake = FakeRoutesClient()
        service = Phase6RouteEstimationService(session, configuration=configuration, model_id="M1", google_client=fake)
        estimate = service.estimate_trip(
            depot=session.get(MasterDepot, "D1"), spbus=[session.get(MasterSPBU, "A")], mt=session.get(MasterMT, "T1"),
            predicted_departure_datetime=datetime(2099, 1, 1, tzinfo=timezone.utc), max_exact_sequence_stops=4,
        )
        assert estimate["routing_mode"] == "DRIVE"
        assert estimate["large_vehicle_used"] is False
        assert estimate["vehicle_profile_snapshot"]["profile_status"] == "NOT_REQUIRED"
        assert estimate["route_geometry_source"] == "GOOGLE_ROUTES_GEOJSON"
        assert estimate["route_geometry"][0] == {"latitude": -6.2, "longitude": 106.84}
        assert estimate["route_geometry"][-1] == {"latitude": -6.2, "longitude": 106.84}
        assert fake.calls == 2
        assert set(fake.routing_modes) == {"DRIVE"}


def test_api_key_is_encrypted_masked_and_never_returned(monkeypatch, caplog) -> None:
    monkeypatch.setenv("GOOGLE_ROUTES_ENCRYPTION_KEY", "test-only-encryption-key-at-least-16")
    get_settings.cache_clear()
    Session = make_session()
    raw_key = "AIzaSyTEST-RAW-KEY-1234567890"
    try:
        with Session() as session:
            seed(session)
            payload = save_google_routes_configuration(session, {"api_key": raw_key}, updated_by="tester")
            configuration = get_google_routes_configuration(session)
            assert configuration.encrypted_api_key != raw_key
            assert raw_key not in configuration.encrypted_api_key
            assert payload["masked_api_key"].endswith("7890")
            assert raw_key not in json.dumps(payload)
            assert raw_key not in caplog.text
            public = public_google_routes_configuration(session)
            assert raw_key not in json.dumps(public)
            assert public["routing_mode"] == "DRIVE"
            assert public["truck_routing_status"] == "DISABLED_FOR_INDONESIA"
            assert "vehicle_profile_readiness" not in public
    finally:
        get_settings.cache_clear()


def test_manual_override_preserves_original_and_recalculates_route() -> None:
    Session = make_session()
    with Session() as session:
        seed(session)
        result = run_prediction(
            session,
            [["LO1", "2026-08-22 09:00:00", "SPBU-A"]],
            [["B1001AA", "2026-08-22 08:00:00"], ["B1002AA", "2026-08-22 08:00:00"]],
        )
        original = result["original_model_prediction"]
        shipment = result["shipments"][0]
        replacement = next(row for row in shipment["candidates"] if row["vehicle_id"] != shipment["assignment"]["assigned_vehicle_id"] and row["compatibility_status"] == "PASS")
        updated = override_assignment(session, result["id"], shipment["assignment"]["id"], replacement["vehicle_id"], "dispatcher", "user")
        assert updated["original_model_prediction"] == original
        assert updated["trips"][0]["vehicle_id"] == replacement["vehicle_id"]
        assert updated["trips"][0]["assignment_status"] == "MANUAL_OVERRIDE"
        assert updated["summary"]["assigned_order_kl"] == 8
        assert updated["hourly_distribution"][-1]["cumulative_kl"] == 8
        assert updated["geographic_routes"]["routes"][0]["vehicle_id"] == replacement["vehicle_id"]
        assert updated["geographic_routes"]["routes"][0]["trip_id"] == updated["trips"][0]["trip_id"]


def test_phase6_has_no_route_optimization_or_full_vrp_client() -> None:
    source = "\n".join(
        Path(path).read_text()
        for path in ("app/google_routes.py", "app/phase6_routing.py", "app/phase6_service.py")
    )
    assert "routeOptimization" not in source
    assert "optimization.googleapis.com" not in source
    assert "def optimize_tours" not in source.lower()
    assert "class VehicleRoutingProblemOptimizer" not in source
